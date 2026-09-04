#!/usr/bin/env python
"""Freeze the 16-panel MetDNA3 external Level-1 benchmark after internal PASS."""
from __future__ import annotations

import argparse
import hashlib
import io
import json
import urllib.parse
import zipfile
from pathlib import Path

import h5py
import numpy as np
import pandas as pd
from openpyxl import load_workbook

from build_bioaware_metdna3_development_manifest import (
    DATASET, WORKBOOK_SUFFIX, decode, fold_membership, sha256, write_json_atomic,
)


SHEETS = {
    "BV2cell": "OrbitrapExploris480_BV2cell",
    "Mouse_brain": "OrbitrapExploris480_MouseBrain",
    "Mouse_liver": "OrbitrapExploris480_MouseLiver",
    "NIST_plasma": "OrbitrapExploris480_NISTplasma",
}
SEPARATION_PANEL = {"hilic": {"HILIC-MS(+)", "HILIC-MS(-)"},
                    "rplc": {"RP-MS(+)", "RP-MS(-)"}}


def stable_split(identities: list[str]) -> pd.DataFrame:
    ordered = sorted(identities, key=lambda value: hashlib.sha256(value.encode()).hexdigest())
    base = {identity: position % 10 for position, identity in enumerate(ordered)}
    return pd.DataFrame([
        {"fold": fold, "ik14": identity, "base_fold": base[identity],
         "role": fold_membership(base[identity], fold)}
        for fold in range(10) for identity in identities
    ])


def read_level1(supplement: Path) -> pd.DataFrame:
    with zipfile.ZipFile(supplement) as bundle:
        matches = [name for name in bundle.namelist() if name.endswith(WORKBOOK_SUFFIX)]
        if len(matches) != 1:
            raise RuntimeError(f"expected one MetDNA3 workbook, got {matches}")
        workbook = load_workbook(io.BytesIO(bundle.read(matches[0])), read_only=True, data_only=True)
    frames: list[pd.DataFrame] = []
    try:
        for sample, sheet in SHEETS.items():
            worksheet = workbook[sheet]
            header = [cell.value for cell in next(worksheet.iter_rows(min_row=2, max_row=2))]
            required = {"LC-MS", "peak_name", "mz", "rt", "name", "formula", "smiles",
                        "inchikey", "adduct", "ms2_score", "iden_type", "confidence_level"}
            if not required.issubset(header):
                raise RuntimeError(f"external sheet schema changed: {sheet}: {required-set(header)}")
            records = []
            for values in worksheet.iter_rows(min_row=3, values_only=True):
                row = dict(zip(header, values))
                if row.get("confidence_level") != "level1":
                    continue
                panel = str(row.get("LC-MS"))
                if panel not in set().union(*SEPARATION_PANEL.values()):
                    continue
                records.append({name: row[name] for name in header})
            frame = pd.DataFrame(records)
            frame["sample_type"] = sample
            frames.append(frame)
    finally:
        workbook.close()
    truth = pd.concat(frames, ignore_index=True)
    truth["ik14"] = truth.inchikey.astype(str).str[:14].str.upper()
    truth["polarity"] = np.where(truth["LC-MS"].astype(str).str.endswith("(+)"),
                                  "positive", "negative")
    truth["separation"] = np.where(truth["LC-MS"].astype(str).str.startswith("HILIC"),
                                    "hilic", "rplc")
    truth["unit_id"] = truth.sample_type.astype(str) + "__" + truth.separation.astype(str)
    truth["panel_id"] = truth.unit_id + "__" + truth.polarity.astype(str)
    if truth.ik14.str.len().ne(14).any() or truth.mz.isna().any() or truth.adduct.isna().any():
        raise RuntimeError("invalid external Level-1 identity/mass/adduct")
    return truth


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--contract", type=Path, default=Path(
        "data/validation/bioaware_metdna3_external_contract_v1/contract.json"))
    parser.add_argument("--router", type=Path, default=Path(
        "data/validation/bioaware_v3_consensus_router_frozen_v2_20260830/artifact.json"))
    parser.add_argument("--internal-result", type=Path, default=Path(
        "data/validation/bioaware_metdna3_internal_rplc_frozen_v3_result_v1/report.json"))
    parser.add_argument("--supplement", type=Path, default=Path(
        "data/external/metdna3_2025/PMC12398597_supplementaryFiles.zip"))
    parser.add_argument("--inventory", type=Path, default=Path(
        "data/external/metdna3_2025/MSV000097913_ccms_peak_inventory.json"))
    parser.add_argument("--reference-hdf5", type=Path, default=Path(
        "data/models/MassSpecGym_MurckoHist_split.hdf5"))
    parser.add_argument("--output-dir", type=Path, default=Path(
        "data/validation/bioaware_metdna3_external_manifest_v1"))
    args = parser.parse_args()
    for path in (args.contract, args.router, args.internal_result, args.supplement,
                 args.inventory, args.reference_hdf5):
        if not path.exists():
            raise FileNotFoundError(path)
    if args.output_dir.exists() and any(args.output_dir.iterdir()):
        raise RuntimeError(f"refusing to overwrite non-empty output: {args.output_dir}")
    contract = json.loads(args.contract.read_text(encoding="utf-8"))
    router = json.loads(args.router.read_text(encoding="utf-8"))
    internal = json.loads(args.internal_result.read_text(encoding="utf-8"))
    if contract.get("status") != "bioaware_metdna3_external_contract_frozen":
        raise RuntimeError("invalid external source contract")
    if router.get("status") != "bioaware_v3_consensus_router_artifact_frozen":
        raise RuntimeError("invalid frozen router")
    if not internal.get("pass_to_external_16_panel"):
        raise RuntimeError("internal RPLC did not unlock external evaluation")
    if internal.get("provenance", {}).get("artifact_sha256") != sha256(args.router):
        raise RuntimeError("internal result belongs to another router")
    if sha256(args.supplement) != contract["supplement"]["sha256"]:
        raise RuntimeError("supplement changed after source lock")

    truth = read_level1(args.supplement)
    if truth.panel_id.nunique() != 16 or truth.unit_id.nunique() != 8:
        raise RuntimeError("external truth does not contain exactly 16 panels / 8 units")
    inventory = json.loads(args.inventory.read_text(encoding="utf-8"))["files"]
    with h5py.File(args.reference_hdf5, "r") as handle:
        ref_ik = np.asarray([decode(value)[:14].upper() for value in handle["INCHIKEY"][:]])
        ref_mz = np.asarray(handle["precursor_mz"][:], dtype=float)
        ref_adduct = np.asarray([decode(value) for value in handle["adduct"][:]])

    args.output_dir.mkdir(parents=True, exist_ok=True)
    unit_reports = {}
    all_downloads = []
    for unit_id, unit_truth in truth.groupby("unit_id", sort=True):
        sample = str(unit_truth.sample_type.iloc[0])
        separation = str(unit_truth.separation.iloc[0])
        folder = f"{sample}_{separation}"
        files = [row.copy() for row in inventory
                 if f"/{folder}/" in row["filepath"]
                 and ("_pos_" in row["filepath"] or "_neg_" in row["filepath"])]
        expected = 15 if unit_id == "Mouse_brain__rplc" else 16
        if len(files) != expected:
            raise RuntimeError(f"{unit_id}: expected {expected} targeted files, got {len(files)}")
        for row in files:
            remote = f"f.{DATASET}/{row['filepath']}"
            row["url"] = "https://massive.ucsd.edu/ProteoSAFe/DownloadResultFile?" + urllib.parse.urlencode({"file": remote})
            row["local_name"] = Path(row["filepath"]).name
            row["unit_id"] = unit_id
        identities = sorted(unit_truth.ik14.unique())
        split = stable_split(identities)
        covered_rows = 0
        covered_ids: set[str] = set()
        ambiguous_ids: set[str] = set()
        for row in unit_truth.itertuples(index=False):
            mask = (ref_adduct == str(row.adduct)) & (
                np.abs(ref_mz - float(row.mz)) <= float(row.mz) * 10e-6)
            candidates = set(ref_ik[mask])
            if row.ik14 in candidates:
                covered_rows += 1
                covered_ids.add(row.ik14)
                if len(candidates) >= 2:
                    ambiguous_ids.add(row.ik14)
        unit_dir = args.output_dir / unit_id
        unit_dir.mkdir()
        truth_path = unit_dir / "external_level1.csv.gz"
        split_path = unit_dir / "identity_splits.csv.gz"
        download_path = unit_dir / "download_manifest.json"
        unit_truth.to_csv(truth_path, index=False, compression="gzip")
        split.to_csv(split_path, index=False, compression="gzip")
        write_json_atomic(download_path, {"dataset": DATASET, "scope": "external", "unit_id": unit_id, "files": files})
        unit_reports[unit_id] = {
            "level1_rows": int(len(unit_truth)),
            "level1_identities": int(len(identities)),
            "panels": unit_truth.panel_id.value_counts().sort_index().to_dict(),
            "targeted_ms2_files": len(files),
            "targeted_ms2_bytes": int(sum(int(row["bytes"]) for row in files)),
            "reference_10ppm_same_adduct": {
                "covered_rows": covered_rows, "covered_identities": len(covered_ids),
                "ambiguous_identities": len(ambiguous_ids),
            },
            "truth_sha256": sha256(truth_path), "split_sha256": sha256(split_path),
            "download_manifest_sha256": sha256(download_path),
        }
        all_downloads.extend(files)
    if len(all_downloads) != 127:
        raise RuntimeError(f"expected 127 external targeted files, got {len(all_downloads)}")
    report = {
        "status": "bioaware_metdna3_external_manifest_frozen",
        "formal": True,
        "units": unit_reports,
        "panels": truth.panel_id.value_counts().sort_index().to_dict(),
        "level1_rows": int(len(truth)),
        "level1_identities": int(truth.ik14.nunique()),
        "targeted_ms2_files": len(all_downloads),
        "targeted_ms2_bytes": int(sum(int(row["bytes"]) for row in all_downloads)),
        "contracts": {
            "router_refit": False, "threshold_tuning": False,
            "one_shot_external_evaluation": True, "phenotype": "forbidden", "P2b": "forbidden",
        },
        "provenance": {
            "source_contract_sha256": sha256(args.contract), "router_sha256": sha256(args.router),
            "internal_result_sha256": sha256(args.internal_result),
            "supplement_sha256": sha256(args.supplement),
            "reference_hdf5_sha256": sha256(args.reference_hdf5),
        },
        "claim_limit": "Frozen external manifest only; no external ranking result.",
    }
    write_json_atomic(args.output_dir / "report.json", report)
    print(json.dumps(report, indent=2))


if __name__ == "__main__":
    main()
