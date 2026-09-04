#!/usr/bin/env python
"""Build the frozen NIST-urine HILIC development manifest after source lock."""
from __future__ import annotations

import argparse
import gzip
import hashlib
import io
import json
import os
import tempfile
import urllib.parse
import zipfile
from pathlib import Path

import h5py
import numpy as np
import pandas as pd
from openpyxl import load_workbook


DATASET = "MSV000097913"
SHEET = "OrbitrapExploris480_NISTurine"
WORKBOOK_SUFFIX = "41467_2025_63536_MOESM3_ESM.xlsx"
DEVELOPMENT_PANELS = {"HILIC-MS(+)", "HILIC-MS(-)"}


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def decode(value: object) -> str:
    if isinstance(value, bytes):
        return value.decode("utf-8", "replace")
    return str(value)


def write_json_atomic(path: Path, payload: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(
        "w", encoding="utf-8", dir=path.parent, delete=False, suffix=".tmp"
    ) as handle:
        json.dump(payload, handle, indent=2, sort_keys=True)
        temporary = Path(handle.name)
    os.replace(temporary, path)


def read_development_level1(supplement: Path) -> pd.DataFrame:
    with zipfile.ZipFile(supplement) as bundle:
        matches = [name for name in bundle.namelist() if name.endswith(WORKBOOK_SUFFIX)]
        if len(matches) != 1:
            raise RuntimeError(f"expected one MetDNA3 workbook, got {matches}")
        workbook = load_workbook(
            io.BytesIO(bundle.read(matches[0])), read_only=True, data_only=True
        )
    try:
        worksheet = workbook[SHEET]
        header = [cell.value for cell in next(worksheet.iter_rows(min_row=2, max_row=2))]
        required = {
            "LC-MS",
            "peak_name",
            "mz",
            "rt",
            "name",
            "formula",
            "smiles",
            "inchikey",
            "adduct",
            "ms2_score",
            "iden_type",
            "confidence_level",
        }
        if not required.issubset(header):
            raise RuntimeError(f"MetDNA3 workbook schema changed: {required-set(header)}")
        records: list[dict] = []
        for values in worksheet.iter_rows(min_row=3, values_only=True):
            row = dict(zip(header, values))
            panel = str(row["LC-MS"])
            if panel not in DEVELOPMENT_PANELS:
                # Stop before the internal-validation RP rows are decoded.
                if panel.startswith("RP-MS"):
                    break
                continue
            if row["confidence_level"] != "level1":
                continue
            records.append({name: row[name] for name in header})
    finally:
        workbook.close()
    frame = pd.DataFrame(records)
    if len(frame) != 751:
        raise RuntimeError(f"expected 751 development Level-1 rows, got {len(frame)}")
    frame["ik14"] = frame["inchikey"].astype(str).str[:14].str.upper()
    frame["polarity"] = np.where(frame["LC-MS"].eq("HILIC-MS(+)"), "positive", "negative")
    return frame


def fold_membership(base: int, fold: int) -> str:
    seed_folds = {fold % 10, (fold + 1) % 10, (fold + 2) % 10}
    return "seed" if base in seed_folds else "heldout"


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--contract",
        type=Path,
        default=Path("data/validation/bioaware_metdna3_external_contract_v1/contract.json"),
    )
    parser.add_argument(
        "--supplement",
        type=Path,
        default=Path("data/external/metdna3_2025/PMC12398597_supplementaryFiles.zip"),
    )
    parser.add_argument(
        "--inventory",
        type=Path,
        default=Path("data/external/metdna3_2025/MSV000097913_ccms_peak_inventory.json"),
    )
    parser.add_argument(
        "--reference-hdf5",
        type=Path,
        default=Path("data/models/MassSpecGym_MurckoHist_split.hdf5"),
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("data/validation/bioaware_metdna3_development_v1"),
    )
    args = parser.parse_args()
    for path in [args.contract, args.supplement, args.inventory, args.reference_hdf5]:
        if not path.exists():
            raise FileNotFoundError(path)
    contract = json.loads(args.contract.read_text(encoding="utf-8"))
    if contract.get("status") != "bioaware_metdna3_external_contract_frozen":
        raise RuntimeError("MetDNA3 source contract is not frozen")
    if contract.get("source_data_opened") is not False:
        raise RuntimeError("source contract must precede workbook inspection")
    if sha256(args.supplement) != contract["supplement"]["sha256"]:
        raise RuntimeError("supplement changed after source lock")
    inventory_payload = json.loads(args.inventory.read_text(encoding="utf-8"))
    inventory = inventory_payload["files"]
    selected = [
        row
        for row in inventory
        if "NIST_urine_hilic/" in row["filepath"]
        and ("_pos_" in row["filepath"] or "_neg_" in row["filepath"])
    ]
    if len(selected) != 16:
        raise RuntimeError(f"expected 16 development targeted-MS2 files, got {len(selected)}")
    for row in selected:
        remote = f"f.{DATASET}/{row['filepath']}"
        row["url"] = (
            "https://massive.ucsd.edu/ProteoSAFe/DownloadResultFile?"
            + urllib.parse.urlencode({"file": remote})
        )
        row["local_name"] = Path(row["filepath"]).name
    truth = read_development_level1(args.supplement)
    identities = sorted(truth["ik14"].unique())
    ordered_identities = sorted(
        identities, key=lambda value: hashlib.sha256(value.encode("utf-8")).hexdigest()
    )
    base_fold = {identity: position % 10 for position, identity in enumerate(ordered_identities)}
    split_rows = [
        {
            "fold": fold,
            "ik14": identity,
            "base_fold": base_fold[identity],
            "role": fold_membership(base_fold[identity], fold),
        }
        for fold in range(10)
        for identity in identities
    ]
    split = pd.DataFrame(split_rows)
    split_counts = split.groupby(["fold", "role"]).size().unstack(fill_value=0)
    if not ((split_counts["seed"] / len(identities)).between(0.27, 0.33)).all():
        raise RuntimeError("deterministic 30/70 split is imbalanced")

    with h5py.File(args.reference_hdf5, "r") as handle:
        ref_ik = np.asarray([decode(value)[:14].upper() for value in handle["INCHIKEY"][:]])
        ref_mz = np.asarray(handle["precursor_mz"][:], dtype=float)
        ref_adduct = np.asarray([decode(value) for value in handle["adduct"][:]])
    covered_rows = 0
    covered_identities: set[str] = set()
    ambiguous_identities: set[str] = set()
    candidate_counts: list[int] = []
    for row in truth.itertuples(index=False):
        mask = (ref_adduct == str(row.adduct)) & (
            np.abs(ref_mz - float(row.mz)) <= float(row.mz) * 10e-6
        )
        candidates = set(ref_ik[mask])
        if row.ik14 in candidates:
            covered_rows += 1
            covered_identities.add(row.ik14)
            candidate_counts.append(len(candidates))
            if len(candidates) >= 2:
                ambiguous_identities.add(row.ik14)
    output = args.output_dir.resolve()
    output.mkdir(parents=True, exist_ok=True)
    truth_path = output / "development_level1.csv.gz"
    split_path = output / "identity_splits.csv.gz"
    download_path = output / "download_manifest.json"
    report_path = output / "report.json"
    if any(path.exists() for path in [truth_path, split_path, download_path, report_path]):
        raise RuntimeError(f"fail-closed: development output already exists: {output}")
    truth.to_csv(truth_path, index=False, compression="gzip")
    split.to_csv(split_path, index=False, compression="gzip")
    write_json_atomic(download_path, {"dataset": DATASET, "files": selected})
    report = {
        "status": "bioaware_metdna3_development_manifest_complete",
        "formal": True,
        "opened_scope": "NIST urine HILIC Level-1 rows only",
        "internal_validation_opened": False,
        "external_test_opened": False,
        "level1_rows": len(truth),
        "level1_identities": len(identities),
        "level1_by_panel": truth["LC-MS"].value_counts().sort_index().to_dict(),
        "reference_10ppm_same_adduct": {
            "covered_rows": covered_rows,
            "covered_identities": len(covered_identities),
            "ambiguous_identities": len(ambiguous_identities),
            "candidate_count_median": float(np.median(candidate_counts)),
            "candidate_count_p90": float(np.quantile(candidate_counts, 0.9)),
        },
        "targeted_ms2_files": len(selected),
        "targeted_ms2_bytes": sum(row["bytes"] for row in selected),
        "split_contract": (
            "10 deterministic identity-level rotations; three hash folds are seed "
            "and seven are held out in every rotation"
        ),
        "provenance": {
            "source_contract_sha256": sha256(args.contract),
            "supplement_sha256": sha256(args.supplement),
            "reference_hdf5_sha256": sha256(args.reference_hdf5),
            "truth_sha256": sha256(truth_path),
            "split_sha256": sha256(split_path),
            "download_manifest_sha256": sha256(download_path),
        },
        "claim_limit": "Development coverage manifest only; no ranking result.",
    }
    write_json_atomic(report_path, report)
    print(json.dumps(report, indent=2), flush=True)


if __name__ == "__main__":
    main()
