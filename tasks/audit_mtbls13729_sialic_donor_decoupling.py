#!/usr/bin/env python
"""Audit same-patient free-Neu5Ac versus activated-donor changes in Rmu.

The source UHPLC-HRMS/MS supplement contains the paired abundance values for
free Neu5Ac, CMP-Neu5Ac and UDP-GlcNAc.  This script reconstructs the Rmu pairs
from the independent sample-metadata supplement, computes patient-level log2
changes, and tests the pre-specified free-pool-minus-donor contrasts.

This is a same-cohort mechanistic decomposition.  It is not independent
replication, flux, or proof of intracellular compartmentalisation.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
from pathlib import Path

import numpy as np
import openpyxl
from scipy import stats


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "data/mtbls13729/source_paper_supplements/pr5c01260_si_005.xlsx"
METADATA = ROOT / "data/mtbls13729/source_paper_supplements/pr5c01260_si_003.xlsx"
DEFAULT_OUTPUT = ROOT / "data/mtbls13729/sialic_donor_decoupling_v1"
TARGETS = [
    {
        "node": "free_neu5ac",
        "name": "N-Acetylneuraminic acid",
        "hmdb": "HMDB0000230",
        "expected_level": "Level 1",
        "role": "free_sialic_acid_pool",
    },
    {
        "node": "cmp_neu5ac",
        "name": "CMP-N-acetylneuraminic acid",
        "hmdb": "HMDB0001176",
        "expected_level": "Level 2",
        "role": "activated_sialic_acid_donor",
    },
    {
        "node": "udp_glcnac",
        "name": "UDP-N-acetylglucosamine",
        "hmdb": "HMDB0000290",
        "expected_level": "Level 1",
        "role": "upstream_nucleotide_sugar_precursor",
    },
]


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1 << 20), b""):
            digest.update(block)
    return digest.hexdigest()


def fail_closed_output(directory: Path) -> None:
    if directory.exists() and any(directory.iterdir()):
        raise RuntimeError(f"refusing to overwrite non-empty output: {directory}")
    directory.mkdir(parents=True, exist_ok=True)


def rmu_pairs(metadata_path: Path) -> list[dict[str, int]]:
    sheet = openpyxl.load_workbook(metadata_path, read_only=True, data_only=True).active
    current_pathology: str | None = None
    tumour_rows: dict[int, dict[str, int | str]] = {}
    normal_tissues: set[int] = set()
    for row in range(3, sheet.max_row + 1):
        tissue = sheet.cell(row, 1).value
        if tissue is None:
            continue
        tissue = int(tissue)
        sample_type = str(sheet.cell(row, 3).value or "")
        pathology = sheet.cell(row, 4).value
        if sample_type == "Tissue-tumor":
            if pathology:
                current_pathology = str(pathology)
            patient = int(sheet.cell(row, 2).value)
            if current_pathology is None:
                raise RuntimeError(f"missing pathology state at tissue {tissue}")
            tumour_rows[tissue] = {
                "patient": patient,
                "pathology": current_pathology,
            }
        elif sample_type == "Tissue-normal":
            normal_tissues.add(tissue)

    pairs: list[dict[str, int]] = []
    for tumour_tissue, metadata in tumour_rows.items():
        if metadata["pathology"] != "Rmu":
            continue
        normal_tissue = tumour_tissue + 1
        if normal_tissue not in normal_tissues:
            raise RuntimeError(f"missing paired normal for tissue {tumour_tissue}")
        pairs.append({
            "patient": int(metadata["patient"]),
            "tumour_tissue": tumour_tissue,
            "normal_tissue": normal_tissue,
        })
    pairs.sort(key=lambda item: item["patient"])
    if [pair["patient"] for pair in pairs] != list(range(21, 31)):
        raise RuntimeError(f"unexpected Rmu patient mapping: {pairs}")
    return pairs


def source_rows(source_path: Path) -> tuple[dict[str, dict[str, object]], dict[int, int]]:
    sheet = openpyxl.load_workbook(source_path, read_only=True, data_only=True)["metabolites"]
    tissue_columns = {
        int(sheet.cell(2, column).value): column
        for column in range(1, sheet.max_column + 1)
        if isinstance(sheet.cell(2, column).value, (int, float))
    }
    if set(tissue_columns) != set(range(1, 61)):
        raise RuntimeError("source abundance table does not contain tissues 1-60 exactly once")

    rows: dict[str, dict[str, object]] = {}
    for target in TARGETS:
        matches = []
        for row in range(3, sheet.max_row + 1):
            if sheet.cell(row, 4).value == target["name"] and sheet.cell(row, 5).value == target["hmdb"]:
                matches.append(row)
        if len(matches) != 1:
            raise RuntimeError(f"expected one row for {target['name']}; found {matches}")
        row = matches[0]
        observed_level = str(sheet.cell(row, 6).value)
        if observed_level != target["expected_level"]:
            raise RuntimeError(f"unexpected identity level for {target['name']}: {observed_level}")
        rows[target["node"]] = {
            **target,
            "row": row,
            "mz": float(sheet.cell(row, 2).value),
            "rt_min": float(sheet.cell(row, 3).value),
            "adduct": str(sheet.cell(row, 13).value),
            "mode": str(sheet.cell(row, 14).value),
            "author_rmu_p": float(sheet.cell(row, 28).value),
            "author_rmu_log2fc": float(sheet.cell(row, 29).value),
            "values": {
                tissue: float(sheet.cell(row, column).value)
                for tissue, column in tissue_columns.items()
            },
        }
    return rows, tissue_columns


def bootstrap_mean_ci(values: np.ndarray, seed: int, resamples: int) -> list[float]:
    rng = np.random.default_rng(seed)
    draws = rng.choice(values, size=(resamples, values.size), replace=True).mean(axis=1)
    return [float(value) for value in np.quantile(draws, [0.025, 0.975])]


def holm_adjust(p_values: list[float]) -> list[float]:
    order = np.argsort(p_values)
    adjusted = np.empty(len(p_values), dtype=float)
    running = 0.0
    total = len(p_values)
    for rank, index in enumerate(order):
        value = min(1.0, (total - rank) * p_values[index])
        running = max(running, value)
        adjusted[index] = running
    return [float(value) for value in adjusted]


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--bootstrap-resamples", type=int, default=100000)
    parser.add_argument("--seed", type=int, default=20260831)
    args = parser.parse_args()
    fail_closed_output(args.output_dir)
    for path in (SOURCE, METADATA):
        if not path.is_file():
            raise FileNotFoundError(path)

    pairs = rmu_pairs(METADATA)
    nodes, _ = source_rows(SOURCE)
    patient_rows: list[dict[str, object]] = []
    deltas: dict[str, np.ndarray] = {}
    for node, source in nodes.items():
        node_deltas: list[float] = []
        values = source["values"]
        for pair in pairs:
            tumour = float(values[pair["tumour_tissue"]])
            normal = float(values[pair["normal_tissue"]])
            if tumour <= 0 or normal <= 0:
                raise RuntimeError(f"non-positive source abundance for {node}/P{pair['patient']}")
            delta = float(np.log2(tumour) - np.log2(normal))
            node_deltas.append(delta)
            patient_rows.append({
                "patient": f"P{pair['patient']}",
                "node": node,
                "metabolite": source["name"],
                "identity_level": source["expected_level"],
                "tumour_tissue": pair["tumour_tissue"],
                "normal_tissue": pair["normal_tissue"],
                "tumour_abundance": tumour,
                "normal_abundance": normal,
                "paired_log2_delta": delta,
            })
        deltas[node] = np.asarray(node_deltas, dtype=float)

    node_summaries: dict[str, object] = {}
    for node, values in deltas.items():
        source = nodes[node]
        ttest = stats.ttest_1samp(values, 0.0)
        wilcoxon = stats.wilcoxon(values, zero_method="wilcox", alternative="two-sided")
        positive = int(np.sum(values > 0))
        sign_test = stats.binomtest(positive, values.size, 0.5, alternative="two-sided")
        node_summaries[node] = {
            "metabolite": source["name"],
            "hmdb": source["hmdb"],
            "identity_level": source["expected_level"],
            "role": source["role"],
            "mz": source["mz"],
            "rt_min": source["rt_min"],
            "adduct": source["adduct"],
            "mode": source["mode"],
            "n_pairs": int(values.size),
            "positive_pairs": positive,
            "mean_paired_log2_delta": float(np.mean(values)),
            "median_paired_log2_delta": float(np.median(values)),
            "bootstrap_mean_95ci": bootstrap_mean_ci(values, args.seed, args.bootstrap_resamples),
            "paired_log_ttest_p": float(ttest.pvalue),
            "wilcoxon_p": float(wilcoxon.pvalue),
            "sign_test_p": float(sign_test.pvalue),
            "author_reported_rmu_log2fc": source["author_rmu_log2fc"],
            "author_reported_rmu_p": source["author_rmu_p"],
        }

    contrast_specs = [
        ("free_neu5ac_minus_cmp_neu5ac", "cmp_neu5ac"),
        ("free_neu5ac_minus_udp_glcnac", "udp_glcnac"),
    ]
    contrast_rows = []
    raw_wilcoxon_p = []
    for name, comparator in contrast_specs:
        values = deltas["free_neu5ac"] - deltas[comparator]
        ttest = stats.ttest_1samp(values, 0.0)
        wilcoxon = stats.wilcoxon(values, zero_method="wilcox", alternative="two-sided")
        positive = int(np.sum(values > 0))
        contrast_rows.append({
            "contrast": name,
            "comparator": comparator,
            "n_pairs": int(values.size),
            "free_pool_change_greater_pairs": positive,
            "mean_log2_delta_difference": float(np.mean(values)),
            "median_log2_delta_difference": float(np.median(values)),
            "bootstrap_mean_95ci": bootstrap_mean_ci(values, args.seed + 1, args.bootstrap_resamples),
            "paired_ttest_p": float(ttest.pvalue),
            "wilcoxon_p": float(wilcoxon.pvalue),
            "sign_test_p": float(stats.binomtest(positive, values.size, 0.5).pvalue),
        })
        raw_wilcoxon_p.append(float(wilcoxon.pvalue))
    for row, adjusted in zip(contrast_rows, holm_adjust(raw_wilcoxon_p)):
        row["wilcoxon_holm_p"] = adjusted

    patient_path = args.output_dir / "rmu_patient_sialic_donor_deltas.csv"
    with patient_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(patient_rows[0]))
        writer.writeheader()
        writer.writerows(patient_rows)

    report = {
        "status": "mtbls13729_sialic_donor_decoupling_audit_complete",
        "formal": True,
        "cohort": "10 Rmu tumour-matched-normal pairs, source HILIC(-) supplement",
        "node_summaries": node_summaries,
        "pre_specified_patient_level_contrasts": contrast_rows,
        "interpretation": {
            "supported": (
                "Free Neu5Ac rises consistently in Rmu while CMP-Neu5Ac and UDP-GlcNAc do not; "
                "the within-patient free-pool change exceeds both donor/precursor changes."
            ),
            "mechanistic_value": (
                "This supplies same-patient evidence for pool-to-activated-donor decoupling and argues "
                "against interpreting free Neu5Ac accumulation as uniform nucleotide-sugar activation."
            ),
            "not_supported": (
                "Static abundance cannot distinguish synthesis, salvage, de-O-acetylation, release, "
                "uptake or reduced glycan incorporation, and CMP-Neu5Ac remains Level 2."
            ),
        },
        "gates": {
            "free_neu5ac_all_pairs_positive": node_summaries["free_neu5ac"]["positive_pairs"] == 10,
            "cmp_neu5ac_not_nominally_increased": node_summaries["cmp_neu5ac"]["paired_log_ttest_p"] >= 0.05,
            "udp_glcnac_not_nominally_increased": node_summaries["udp_glcnac"]["paired_log_ttest_p"] >= 0.05,
            "both_contrast_bootstrap_lower_bounds_positive": all(
                row["bootstrap_mean_95ci"][0] > 0 for row in contrast_rows
            ),
            "both_contrast_wilcoxon_holm_p_le_0_05": all(
                row["wilcoxon_holm_p"] <= 0.05 for row in contrast_rows
            ),
        },
        "claim_limit": (
            "Same-cohort paired abundance decomposition. It is not independent replication, subcellular "
            "localisation, glycan destination, isotope flux, enzyme activity or causality."
        ),
        "parameters": {
            "bootstrap_resamples": args.bootstrap_resamples,
            "seed": args.seed,
            "planned_comparators": [item[1] for item in contrast_specs],
        },
        "provenance": {
            "source_metabolite_supplement_sha256": sha256(SOURCE),
            "source_sample_metadata_sha256": sha256(METADATA),
            "patient_csv_sha256": sha256(patient_path),
            "script_sha256": sha256(Path(__file__)),
        },
    }
    if not all(report["gates"].values()):
        raise RuntimeError(f"sialic donor decoupling gates failed: {report['gates']}")
    report_path = args.output_dir / "report.json"
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()

