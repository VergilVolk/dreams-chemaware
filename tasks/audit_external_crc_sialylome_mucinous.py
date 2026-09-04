#!/usr/bin/env python
"""Audit the external CRC sialyltransferase-score/mucinous association.

The source publication integrates TCGA, Sidra-LUMC and CPTAC-2 transcriptomes.
Its result is useful glycobiology context, but it is neither metabolomics nor
an independent Neu5Ac abundance replication.  This script freezes the public
supplement and reconstructs only the reported histology 2x2 table.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import math
from datetime import datetime, timezone
from pathlib import Path
from zipfile import ZipFile

import openpyxl
import requests
from scipy.stats import fisher_exact


ROOT = Path(__file__).resolve().parents[1]
SOURCE_URL = (
    "https://mdpi-res.com/d_attachment/biology/biology-15-00705/"
    "article_deploy/biology-15-00705-s001.zip"
)
EXPECTED_ZIP_SHA256 = "4e2b5098fc07df3c37b7a3f505ecbc902e191769f2dd95e4882dd58310aff0f7"
DEFAULT_OUTPUT = ROOT / "data/external/CRC_sialylome_mucinous_Biology2026_20260831"
EXPECTED_GENES = [
    "ST3GAL1", "ST3GAL2", "ST3GAL3", "ST3GAL4", "ST3GAL5", "ST3GAL6",
    "ST6GAL1", "ST6GAL2", "ST6GALNAC1", "ST6GALNAC2", "ST6GALNAC3",
    "ST6GALNAC4", "ST6GALNAC5", "ST6GALNAC6", "ST8SIA1", "ST8SIA2",
    "ST8SIA3", "ST8SIA4", "ST8SIA5", "ST8SIA6",
]


def sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1 << 20), b""):
            digest.update(block)
    return digest.hexdigest()


def fail_closed_output(directory: Path) -> None:
    if directory.exists() and any(directory.iterdir()):
        raise RuntimeError(f"refusing to overwrite non-empty output: {directory}")
    directory.mkdir(parents=True, exist_ok=True)


def locate_histology_rows(sheet: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, list[object]]:
    rows = [[sheet.cell(row, column).value for column in range(1, 11)] for row in range(1, sheet.max_row + 1)]
    feature_rows = [i for i, row in enumerate(rows) if row[0] == "Histology"]
    if len(feature_rows) != 1:
        raise RuntimeError(f"expected one Histology row; found {len(feature_rows)}")
    index = feature_rows[0]
    block = rows[index:index + 3]
    if [block[1][1], block[2][1]] != ["Mucinous Adenocarcinoma", "Non-mucinous Adenocarcinoma"]:
        raise RuntimeError(f"unexpected histology categories: {block}")
    return {"summary": block[0], "mucinous": block[1], "non_mucinous": block[2]}


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--timeout", type=float, default=60.0)
    args = parser.parse_args()
    fail_closed_output(args.output_dir)

    response = requests.get(
        SOURCE_URL,
        headers={"User-Agent": "DreaMS-external-evidence-audit/1.0"},
        timeout=args.timeout,
    )
    response.raise_for_status()
    archive_bytes = response.content
    observed_hash = sha256_bytes(archive_bytes)
    if observed_hash != EXPECTED_ZIP_SHA256:
        raise RuntimeError(f"supplement hash mismatch: {observed_hash}")
    archive_path = args.output_dir / "biology-15-00705-s001.zip"
    archive_path.write_bytes(archive_bytes)

    with ZipFile(archive_path) as archive:
        names = archive.namelist()
        if names != ["biology-4217529-supplementary.xlsx"]:
            raise RuntimeError(f"unexpected archive contents: {names}")
        workbook_path = args.output_dir / names[0]
        workbook_path.write_bytes(archive.read(names[0]))

    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    expected_sheets = [f"Supplementary Materials {i}" for i in range(1, 12)]
    if workbook.sheetnames != expected_sheets:
        raise RuntimeError(f"unexpected workbook sheets: {workbook.sheetnames}")

    gene_sheet = workbook["Supplementary Materials 1"]
    genes = [gene_sheet.cell(row, 1).value for row in range(3, gene_sheet.max_row + 1)]
    genes = [str(gene) for gene in genes if gene]
    if genes != EXPECTED_GENES:
        raise RuntimeError(f"unexpected sialylome score genes: {genes}")

    histology = locate_histology_rows(workbook["Supplementary Materials 3"])
    summary = histology["summary"]
    mucinous = histology["mucinous"]
    non_mucinous = histology["non_mucinous"]
    counts = {
        "mucinous_low": int(mucinous[2]),
        "mucinous_high": int(mucinous[3]),
        "mucinous_total": int(mucinous[4]),
        "non_mucinous_low": int(non_mucinous[2]),
        "non_mucinous_high": int(non_mucinous[3]),
        "non_mucinous_total": int(non_mucinous[4]),
        "histology_total": int(summary[4]),
    }
    if counts["mucinous_low"] + counts["mucinous_high"] != counts["mucinous_total"]:
        raise RuntimeError("mucinous counts do not sum")
    if counts["non_mucinous_low"] + counts["non_mucinous_high"] != counts["non_mucinous_total"]:
        raise RuntimeError("non-mucinous counts do not sum")
    if counts["mucinous_total"] + counts["non_mucinous_total"] != counts["histology_total"]:
        raise RuntimeError("histology counts do not sum")

    table = [
        [counts["mucinous_high"], counts["mucinous_low"]],
        [counts["non_mucinous_high"], counts["non_mucinous_low"]],
    ]
    fisher = fisher_exact(table, alternative="two-sided")
    standard_error = math.sqrt(sum(1 / value for row in table for value in row))
    log_or = math.log(float(fisher.statistic))
    ci95 = [math.exp(log_or - 1.96 * standard_error), math.exp(log_or + 1.96 * standard_error)]

    report = {
        "status": "external_crc_sialylome_mucinous_audit_complete",
        "formal": True,
        "retrieved_at_utc": datetime.now(timezone.utc).isoformat(),
        "paper": {
            "title": "The Colorectal Cancer Glycocode: Tumour Sialylation Is Associated with an Immune-Excluded Phenotype and Distinct Therapeutic Signatures",
            "doi": "10.3390/biology15090705",
            "pmid": "42117844",
            "cohorts": ["TCGA", "Sidra-LUMC", "CPTAC-2"],
            "integrated_n": 988,
        },
        "score_definition": {
            "name": "Sialylome Activity score",
            "genes": genes,
            "gene_count": len(genes),
            "interpretation": (
                "The supplement defines the score using 20 sialyltransferase genes. "
                "It is not a measured glycan, free Neu5Ac abundance, flux, or enzyme-activity score."
            ),
        },
        "histology_table": {
            **counts,
            "published_chi_square": float(summary[5]),
            "published_df": int(summary[6]),
            "published_p": str(summary[7]),
            "published_fdr_q": str(summary[8]),
            "mucinous_high_fraction": counts["mucinous_high"] / counts["mucinous_total"],
            "non_mucinous_high_fraction": counts["non_mucinous_high"] / counts["non_mucinous_total"],
            "reconstructed_odds_ratio": float(fisher.statistic),
            "reconstructed_odds_ratio_ci95": ci95,
            "reconstructed_fisher_exact_p": float(fisher.pvalue),
        },
        "overlap_audit": {
            "independent_of_local_mtbls13729": True,
            "independent_of_local_tcga_branch_analysis": False,
            "reason": (
                "The integrated publication includes TCGA-COADREAD, already used by the local transcriptomic "
                "branch audit. Sidra-LUMC and CPTAC-2 are additional cohorts, but the supplement does not "
                "report the mucinous association separately by cohort."
            ),
        },
        "interpretation": {
            "supported": (
                "Mucinous histology is enriched in the integrated high-sialyltransferase-expression group "
                "(55.2% versus 28.8%; reconstructed OR about 3.04)."
            ),
            "relevance": (
                "This independently supports prioritising branch-aware sialylation biology in mucinous CRC "
                "and is compatible with the local hybrid-glycome model."
            ),
            "not_supported": (
                "It does not replicate feature703, free Neu5Ac abundance, a specific glycan linkage, or the "
                "MTBLS13729 Rmu-versus-RN effect."
            ),
        },
        "claim_limit": (
            "External pooled transcriptomic context with partial TCGA overlap; not an independent metabolite "
            "replication and not cohort-specific validation of the local branch model."
        ),
        "provenance": {
            "source_url": SOURCE_URL,
            "archive_sha256": sha256_file(archive_path),
            "workbook_sha256": sha256_file(workbook_path),
            "script_sha256": sha256_file(Path(__file__)),
        },
    }
    report_path = args.output_dir / "report.json"
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()

