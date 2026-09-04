#!/usr/bin/env python
"""Open and freeze only the preregistered NIST-urine RPLC internal panels."""
from __future__ import annotations

import argparse
import hashlib
import io
import json
from pathlib import Path
import urllib.parse
import zipfile

import h5py
import numpy as np
import pandas as pd
from openpyxl import load_workbook

from build_bioaware_metdna3_development_manifest import (
    DATASET, SHEET, WORKBOOK_SUFFIX, decode, fold_membership, sha256,
    write_json_atomic,
)


PANELS = {"RP-MS(+)", "RP-MS(-)"}


def read_internal_level1(supplement: Path) -> pd.DataFrame:
    with zipfile.ZipFile(supplement) as bundle:
        matches = [name for name in bundle.namelist() if name.endswith(WORKBOOK_SUFFIX)]
        if len(matches) != 1:
            raise RuntimeError(f"expected one MetDNA3 workbook, got {matches}")
        workbook = load_workbook(io.BytesIO(bundle.read(matches[0])), read_only=True, data_only=True)
    try:
        worksheet = workbook[SHEET]
        header = [cell.value for cell in next(worksheet.iter_rows(min_row=2, max_row=2))]
        required = {"LC-MS", "peak_name", "mz", "rt", "name", "formula", "smiles",
                    "inchikey", "adduct", "ms2_score", "iden_type", "confidence_level"}
        if not required.issubset(header):
            raise RuntimeError(f"MetDNA3 workbook schema changed: {required-set(header)}")
        records = []
        for values in worksheet.iter_rows(min_row=3, values_only=True):
            row = dict(zip(header, values))
            if str(row.get("LC-MS")) not in PANELS or row.get("confidence_level") != "level1":
                continue
            records.append({name: row[name] for name in header})
    finally:
        workbook.close()
    frame = pd.DataFrame(records)
    if frame.empty:
        raise RuntimeError("no NIST-urine RPLC Level-1 rows")
    frame["ik14"] = frame.inchikey.astype(str).str[:14].str.upper()
    frame["polarity"] = np.where(frame["LC-MS"].eq("RP-MS(+)"), "positive", "negative")
    if frame.ik14.str.len().ne(14).any() or frame.mz.isna().any() or frame.adduct.isna().any():
        raise RuntimeError("invalid internal Level-1 identity/mass/adduct")
    return frame


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--contract", type=Path, default=Path(
        "data/validation/bioaware_metdna3_external_contract_v1/contract.json"))
    parser.add_argument("--router", type=Path, default=Path(
        "data/validation/bioaware_v3_consensus_router_frozen_v2_20260830/artifact.json"))
    parser.add_argument("--supplement", type=Path, default=Path(
        "data/external/metdna3_2025/PMC12398597_supplementaryFiles.zip"))
    parser.add_argument("--inventory", type=Path, default=Path(
        "data/external/metdna3_2025/MSV000097913_ccms_peak_inventory.json"))
    parser.add_argument("--reference-hdf5", type=Path, default=Path(
        "data/models/MassSpecGym_MurckoHist_split.hdf5"))
    parser.add_argument("--output-dir", type=Path, default=Path(
        "data/validation/bioaware_metdna3_internal_rplc_manifest_v1"))
    args = parser.parse_args()
    for path in (args.contract, args.router, args.supplement, args.inventory, args.reference_hdf5):
        if not path.exists():
            raise FileNotFoundError(path)
    if args.output_dir.exists() and any(args.output_dir.iterdir()):
        raise RuntimeError(f"refusing to overwrite non-empty output: {args.output_dir}")
    contract = json.loads(args.contract.read_text(encoding="utf-8"))
    router = json.loads(args.router.read_text(encoding="utf-8"))
    if contract.get("status") != "bioaware_metdna3_external_contract_frozen":
        raise RuntimeError("invalid external source contract")
    if not router.get("gates", {}).get("pass_to_frozen_rplc_internal_validation"):
        raise RuntimeError("frozen router did not unlock RPLC internal validation")
    if not router.get("contracts", {}).get("evaluation_must_load_this_artifact_without_refit"):
        raise RuntimeError("router is not a frozen evaluation artifact")
    if sha256(args.supplement) != contract["supplement"]["sha256"]:
        raise RuntimeError("supplement changed after source lock")
    truth = read_internal_level1(args.supplement)
    identities = sorted(truth.ik14.unique())
    ordered = sorted(identities, key=lambda value: hashlib.sha256(value.encode()).hexdigest())
    base_fold = {identity: position % 10 for position, identity in enumerate(ordered)}
    split = pd.DataFrame([
        {"fold": fold, "ik14": identity, "base_fold": base_fold[identity],
         "role": fold_membership(base_fold[identity], fold)}
        for fold in range(10) for identity in identities
    ])
    inventory = json.loads(args.inventory.read_text(encoding="utf-8"))["files"]
    selected = [row.copy() for row in inventory if "NIST_urine_rplc/" in row["filepath"]
                and ("_pos_" in row["filepath"] or "_neg_" in row["filepath"])]
    if len(selected) != 16:
        raise RuntimeError(f"expected 16 NIST-urine RPLC files, got {len(selected)}")
    for row in selected:
        remote = f"f.{DATASET}/{row['filepath']}"
        row["url"] = "https://massive.ucsd.edu/ProteoSAFe/DownloadResultFile?" + urllib.parse.urlencode({"file": remote})
        row["local_name"] = Path(row["filepath"]).name

    with h5py.File(args.reference_hdf5, "r") as handle:
        ref_ik = np.asarray([decode(value)[:14].upper() for value in handle["INCHIKEY"][:]])
        ref_mz = np.asarray(handle["precursor_mz"][:], dtype=float)
        ref_adduct = np.asarray([decode(value) for value in handle["adduct"][:]])
    covered_rows = 0
    covered_ids = set()
    ambiguous_ids = set()
    candidate_counts = []
    for row in truth.itertuples(index=False):
        mask = (ref_adduct == str(row.adduct)) & (np.abs(ref_mz-float(row.mz)) <= float(row.mz)*10e-6)
        candidates = set(ref_ik[mask])
        if row.ik14 in candidates:
            covered_rows += 1
            covered_ids.add(row.ik14)
            candidate_counts.append(len(candidates))
            if len(candidates) >= 2:
                ambiguous_ids.add(row.ik14)
    args.output_dir.mkdir(parents=True)
    truth_path = args.output_dir / "internal_level1.csv.gz"
    split_path = args.output_dir / "identity_splits.csv.gz"
    download_path = args.output_dir / "download_manifest.json"
    truth.to_csv(truth_path, index=False)
    split.to_csv(split_path, index=False)
    write_json_atomic(download_path, {"dataset": DATASET, "files": selected})
    report = {
        "status": "bioaware_metdna3_internal_rplc_manifest_complete",
        "formal": True,
        "opened_scope": "NIST urine RPLC positive and negative Level-1 rows only",
        "external_16_panel_outcomes_opened": False,
        "level1_rows": int(len(truth)),
        "level1_identities": int(len(identities)),
        "level1_by_panel": truth["LC-MS"].value_counts().sort_index().to_dict(),
        "reference_10ppm_same_adduct": {
            "covered_rows": covered_rows, "covered_identities": len(covered_ids),
            "ambiguous_identities": len(ambiguous_ids),
            "candidate_count_median": float(np.median(candidate_counts)) if candidate_counts else 0.0,
        },
        "targeted_ms2_files": len(selected),
        "contracts": {
            "router_refit_on_rplc": False,
            "threshold_tuning_on_rplc": False,
            "phenotype": "forbidden",
            "P2b": "forbidden",
        },
        "provenance": {
            "source_contract_sha256": sha256(args.contract),
            "router_artifact_sha256": sha256(args.router),
            "supplement_sha256": sha256(args.supplement),
            "reference_hdf5_sha256": sha256(args.reference_hdf5),
            "truth_sha256": sha256(truth_path),
            "split_sha256": sha256(split_path),
            "download_manifest_sha256": sha256(download_path),
        },
        "claim_limit": "Internal validation manifest only; no ranking result and no external SOTA claim.",
    }
    write_json_atomic(args.output_dir / "report.json", report)
    print(json.dumps(report, indent=2))


if __name__ == "__main__":
    main()
